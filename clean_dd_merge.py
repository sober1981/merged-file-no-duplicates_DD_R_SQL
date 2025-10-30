#!/usr/bin/env python3
"""
CLEAN DD MERGE - Directional Duplicates Removal Script
Version: 1.0
Date: 2025-10-29

This script removes duplicate DIRECTIONAL runs from the merged scorecard data.

DUPLICATE DETECTION & REMOVAL LOGIC:
1. Duplicates are identified using THREE criteria (all must match):
   - JOB_NUM (must match exactly)
   - Total Hrs within ±5 hours tolerance
   - Last 3 digits of Serial Number (SN) must match

2. Reference files are identified by SOURCE:
   - Motor KPI (SOURCE='Motor_KPI'): Reference for ALL Directional job types
   - CAM Run Tracker (SOURCE='CAM_Run_Tracker'): Reference for ALL Rental job types

3. DUPLICATE REMOVAL (KEY DIFFERENCE):
   - DIRECTIONAL duplicates: POG rows matching Motor KPI are REMOVED
   - RENTAL duplicates: POG rows matching CAM Run Tracker are KEPT (not removed)

4. POG files processing:
   - If POG JOB_TYPE = "Directional" -> Check against Motor KPI, REMOVE if duplicate
   - If POG JOB_TYPE = "Rental" -> Check against CAM Run Tracker, KEEP even if duplicate

DATA CLEANING:
- Rows with both Total Hrs = 0/blank AND TOTAL_DRILL = 0/blank are removed
- Directional duplicate rows are REMOVED from output
- Rental duplicate rows are KEPT in output
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import glob
import os

# Configuration
TOTAL_HRS_TOLERANCE = 5  # ±5 hours tolerance
SN_LAST_DIGITS = 3       # Match last 3 digits of Serial Number

# Yellow fill for highlighting Rental duplicates
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def find_merged_file():
    """Find the most recent MERGED_DATA file in the current directory."""
    print("\nSearching for merged data file...")
    pattern = "MERGED_DATA*.xlsx"
    matches = glob.glob(pattern)

    if len(matches) == 0:
        print(f"  ERROR: No file found matching pattern '{pattern}'")
        return None
    elif len(matches) > 1:
        # Sort by modification time, get most recent
        matches.sort(key=os.path.getmtime, reverse=True)
        print(f"  WARNING: Multiple files found. Using most recent: {matches[0]}")
    else:
        print(f"  Found: {matches[0]}")

    return matches[0]


def extract_last_digits(sn, num_digits=3):
    """
    Extract the last N digits from a serial number.
    Handles various formats and extracts only numeric digits.

    Args:
        sn: Serial number (string or number)
        num_digits: Number of digits to extract from the end

    Returns:
        String of last N digits, or empty string if not enough digits
    """
    if pd.isna(sn):
        return ""

    # Convert to string and extract only digits
    sn_str = str(sn)
    digits_only = ''.join(filter(str.isdigit, sn_str))

    # Return last N digits
    if len(digits_only) >= num_digits:
        return digits_only[-num_digits:]
    else:
        return digits_only  # Return whatever digits we have


def is_duplicate(row, reference_df, job_type):
    """
    Check if a row is a duplicate based on:
    1. JOB_NUM matches
    2. Total Hrs within tolerance (±5 hours) OR TOTAL_DRILL within tolerance if hrs blank
    3. Last 3 digits of SN match

    Args:
        row: The row to check (from POG files)
        reference_df: DataFrame to check against (Motor KPI or CAM Run Tracker)
        job_type: "Directional" or "Rental"

    Returns:
        Boolean indicating if row is a duplicate
    """
    # Get values from the row being checked
    job_num = row['JOB_NUM']
    total_hrs = row['Total Hrs (C+D)']
    total_drill = row['TOTAL_DRILL']
    sn_last_digits = row['SN_LAST_3']

    # Skip if missing critical identifiers
    if pd.isna(job_num) or job_num == '':
        return False

    if pd.isna(total_hrs):
        total_hrs = 0

    if pd.isna(total_drill):
        total_drill = 0

    if sn_last_digits == '':
        # If no SN digits, can't match on third criterion
        # But we can still check JOB_NUM + Total Hrs/Drill
        pass

    # Filter reference data by JOB_NUM first (must match exactly)
    matching_jobs = reference_df[reference_df['JOB_NUM'] == job_num]

    if len(matching_jobs) == 0:
        return False  # No matching job number

    # NEW LOGIC: Check if POG row represents multiple reference runs combined
    # Group reference rows by SN (same JOB_NUM + same SN last 3 digits)
    if sn_last_digits != '':
        matching_sn_rows = matching_jobs[matching_jobs['SN_LAST_3'] == sn_last_digits]

        if len(matching_sn_rows) > 0:
            # Sum all Total Hrs for matching JOB_NUM + SN combinations
            sum_ref_hours = matching_sn_rows['Total Hrs (C+D)'].fillna(0).sum()

            # Check if POG total hours matches the SUM of reference hours
            if total_hrs > 0 and sum_ref_hours > 0:
                if abs(total_hrs - sum_ref_hours) <= TOTAL_HRS_TOLERANCE:
                    return True  # POG row represents combined reference runs
            else:
                # If Total Hrs is blank/zero, check TOTAL_DRILL instead
                sum_ref_drill = matching_sn_rows['TOTAL_DRILL'].fillna(0).sum()
                if total_drill > 0 and sum_ref_drill > 0:
                    if abs(total_drill - sum_ref_drill) <= TOTAL_HRS_TOLERANCE:
                        return True  # POG row represents combined reference runs (matched by drill)

    # ORIGINAL LOGIC: Check each matching job for Total Hrs tolerance and SN match
    for _, ref_row in matching_jobs.iterrows():
        ref_total_hrs = ref_row['Total Hrs (C+D)']
        ref_total_drill = ref_row['TOTAL_DRILL']
        ref_sn_last_digits = ref_row['SN_LAST_3']

        if pd.isna(ref_total_hrs):
            ref_total_hrs = 0

        if pd.isna(ref_total_drill):
            ref_total_drill = 0

        # Check if total hours are within tolerance
        # If both Total Hrs are blank/zero, fall back to TOTAL_DRILL comparison
        hrs_match = False
        if total_hrs > 0 or ref_total_hrs > 0:
            hrs_match = abs(total_hrs - ref_total_hrs) <= TOTAL_HRS_TOLERANCE
        else:
            # Both Total Hrs are blank/zero, check TOTAL_DRILL instead
            if total_drill > 0 or ref_total_drill > 0:
                hrs_match = abs(total_drill - ref_total_drill) <= TOTAL_HRS_TOLERANCE

        # Check if SN last digits match
        sn_match = False
        if sn_last_digits != '' and ref_sn_last_digits != '':
            sn_match = (sn_last_digits == ref_sn_last_digits)
        elif sn_last_digits == '' and ref_sn_last_digits == '':
            # Both have no SN - consider it a match if other criteria met
            sn_match = True

        # If both hours/drill and SN match, it's a duplicate
        if hrs_match and sn_match:
            return True

    return False


def remove_empty_runs(df):
    """
    Remove rows where both Total Hrs = 0/blank AND TOTAL_DRILL = 0/blank.
    Keep rows if at least one has a value.

    Args:
        df: DataFrame with merged data

    Returns:
        Filtered DataFrame and count of removed rows
    """
    print("\nStep 1: Removing runs with no hours and no drill distance...")

    initial_count = len(df)

    # Create conditions for empty/zero values
    total_hrs_empty = (df['Total Hrs (C+D)'].isna()) | (df['Total Hrs (C+D)'] == 0)
    total_drill_empty = (df['TOTAL_DRILL'].isna()) | (df['TOTAL_DRILL'] == 0)

    # Keep rows where at least one is NOT empty/zero
    # IMPORTANT: Don't use .copy() to preserve original indices
    df_filtered = df[~(total_hrs_empty & total_drill_empty)]

    removed_count = initial_count - len(df_filtered)

    print(f"  Removed {removed_count} rows with no hours and no drill distance")
    print(f"  Remaining rows: {len(df_filtered)}")

    return df_filtered, removed_count


def detect_duplicates(df):
    """
    Main function to detect duplicates in the merged data.

    Args:
        df: DataFrame with merged data

    Returns:
        DataFrame with 'IS_DUPLICATE' column added
    """
    print("\nStep 2: Detecting duplicates...")

    # Add column to extract last 3 digits of SN
    print("  Extracting last 3 digits from Serial Numbers...")
    df = df.copy()  # Make an explicit copy to avoid SettingWithCopyWarning
    df['SN_LAST_3'] = df['SN'].apply(lambda x: extract_last_digits(x, SN_LAST_DIGITS))

    # Add columns to mark duplicates
    df['IS_DUPLICATE'] = False  # For Directional duplicates (to be removed)
    df['IS_RENTAL_DUPLICATE'] = False  # For Rental duplicates (to be highlighted)

    # Separate by SOURCE (primary identifier for reference files)
    # Reference files are identified by SOURCE, not JOB_TYPE:
    #   - All Motor KPI rows are Directional reference (regardless of JOB_TYPE)
    #   - All CAM Run Tracker rows are Rental reference (regardless of JOB_TYPE)
    motor_kpi = df[df['SOURCE'] == 'Motor_KPI'].copy()
    cam_tracker = df[df['SOURCE'] == 'CAM_Run_Tracker'].copy()
    pog_cam = df[df['SOURCE'] == 'POG_CAM_Usage'].copy()
    pog_mm = df[df['SOURCE'] == 'POG_MM_Usage'].copy()

    print(f"\n  Source breakdown:")
    print(f"    Motor KPI (Directional reference - by SOURCE): {len(motor_kpi)} rows")
    print(f"    CAM Run Tracker (Rental reference - by SOURCE): {len(cam_tracker)} rows")
    print(f"    POG CAM Usage (checked by JOB_TYPE): {len(pog_cam)} rows")
    print(f"    POG MM Usage (checked by JOB_TYPE): {len(pog_mm)} rows")

    # Check POG files for duplicates against reference files
    # KEY DIFFERENCE: Only mark DIRECTIONAL duplicates for removal
    # RENTAL duplicates are NOT marked (will be kept in output)
    directional_duplicate_count = 0
    rental_duplicate_count = 0

    # Check POG CAM: Use JOB_TYPE to determine reference
    print("\n  Checking POG CAM for duplicates...")
    for idx, row in pog_cam.iterrows():
        job_type = row.get('JOB_TYPE', '')

        if job_type == 'Directional':
            # POG is Directional -> Check against Motor KPI, MARK FOR REMOVAL if duplicate
            if is_duplicate(row, motor_kpi, 'Directional'):
                df.at[idx, 'IS_DUPLICATE'] = True
                directional_duplicate_count += 1
        elif job_type == 'Rental':
            # POG is Rental -> Check against CAM Run Tracker, MARK FOR HIGHLIGHTING (but keep)
            if is_duplicate(row, cam_tracker, 'Rental'):
                df.at[idx, 'IS_RENTAL_DUPLICATE'] = True
                rental_duplicate_count += 1

    # Check POG MM: Use JOB_TYPE to determine reference
    print("  Checking POG MM for duplicates...")
    for idx, row in pog_mm.iterrows():
        job_type = row.get('JOB_TYPE', '')

        if job_type == 'Directional':
            # POG is Directional -> Check against Motor KPI, MARK FOR REMOVAL if duplicate
            if is_duplicate(row, motor_kpi, 'Directional'):
                df.at[idx, 'IS_DUPLICATE'] = True
                directional_duplicate_count += 1
        elif job_type == 'Rental':
            # POG is Rental -> Check against CAM Run Tracker, MARK FOR HIGHLIGHTING (but keep)
            if is_duplicate(row, cam_tracker, 'Rental'):
                df.at[idx, 'IS_RENTAL_DUPLICATE'] = True
                rental_duplicate_count += 1

    print(f"\n  Directional duplicates (will be REMOVED): {directional_duplicate_count}")
    print(f"  Rental duplicates (will be KEPT): {rental_duplicate_count}")
    print(f"  Total duplicates detected: {directional_duplicate_count + rental_duplicate_count}")

    return df, rental_duplicate_count


def remove_directional_duplicates(df):
    """
    Remove rows marked as duplicates (Directional duplicates only).

    Args:
        df: DataFrame with IS_DUPLICATE column

    Returns:
        DataFrame with duplicates removed, count of removed rows
    """
    print("\nStep 3: Removing Directional duplicate rows...")

    initial_count = len(df)
    df_clean = df[df['IS_DUPLICATE'] == False].copy()
    removed_count = initial_count - len(df_clean)

    print(f"  Removed {removed_count} Directional duplicate rows")
    print(f"  Remaining rows: {len(df_clean)}")

    return df_clean, removed_count


def highlight_rental_duplicates(file_path, df):
    """
    Highlight Rental duplicate rows in yellow in the Excel file.

    Args:
        file_path: Path to the output Excel file
        df: DataFrame with IS_RENTAL_DUPLICATE column
    """
    print("\nStep 4: Highlighting Rental duplicate rows in yellow...")

    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active

    # Get Rental duplicate rows and their POSITION (not index) in the DataFrame
    # When exported to Excel with index=False, position determines Excel row number
    df_reset = df.reset_index(drop=True)
    rental_dup_positions = df_reset[df_reset['IS_RENTAL_DUPLICATE'] == True].index.tolist()

    highlighted_count = 0
    for position in rental_dup_positions:
        excel_row = position + 2  # +2 because Excel is 1-indexed and has header row

        # Highlight all cells in the row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = YELLOW_FILL

        highlighted_count += 1

    # Save the workbook
    wb.save(file_path)

    print(f"  Highlighted {highlighted_count} Rental duplicate rows in yellow")


def generate_summary_report(original_count, after_empty_removal, directional_dup_count, rental_dup_count, final_count, output_file):
    """
    Generate a summary report of the cleaning process.

    Args:
        original_count: Original number of rows
        after_empty_removal: Number of rows after removing empty runs
        directional_dup_count: Number of Directional duplicates removed
        rental_dup_count: Number of Rental duplicates detected (but kept)
        final_count: Final number of rows in output
        output_file: Name of output file
    """
    removed_empty = original_count - after_empty_removal

    print("\n" + "="*70)
    print("CLEAN DD MERGE - SUMMARY")
    print("="*70)
    print(f"\nOriginal merged file rows:                  {original_count}")
    print(f"Rows removed (no hrs & no drill):           {removed_empty}")
    print(f"Rows after empty removal:                   {after_empty_removal}")
    print(f"\nDirectional duplicates (REMOVED):           {directional_dup_count}")
    print(f"Rental duplicates (KEPT):                   {rental_dup_count}")
    print(f"Total duplicates detected:                  {directional_dup_count + rental_dup_count}")
    print(f"\nFinal row count in output:                  {final_count}")
    print(f"Clean rows (no duplicates):                 {final_count - rental_dup_count}")
    print(f"\nOutput file: {output_file}")
    print("\nNOTE: Directional duplicates have been REMOVED from the output.")
    print("      Rental duplicates have been KEPT and HIGHLIGHTED in YELLOW.")
    print("="*70)


def main():
    """Main execution function."""
    print("="*70)
    print("CLEAN DD MERGE - Directional Duplicates Removal")
    print("="*70)
    print("\nCriteria for duplicate detection:")
    print(f"  1. JOB_NUM must match exactly")
    print(f"  2. Total Hrs within ±{TOTAL_HRS_TOLERANCE} hours")
    print(f"  3. Last {SN_LAST_DIGITS} digits of Serial Number must match")
    print("\nReference files (identified by SOURCE):")
    print("  - Motor KPI (SOURCE='Motor_KPI'): Reference for ALL Directional runs")
    print("  - CAM Run Tracker (SOURCE='CAM_Run_Tracker'): Reference for ALL Rental runs")
    print("\nDuplicate Removal Logic:")
    print("  - Directional POG duplicates (vs Motor KPI) -> REMOVED")
    print("  - Rental POG duplicates (vs CAM Run Tracker) -> KEPT")

    # Find merged file
    merged_file = find_merged_file()
    if merged_file is None:
        print("\nERROR: No merged file found. Please run merge script first.")
        return

    # Read merged data
    print(f"\nReading merged data from: {merged_file}")
    df = pd.read_excel(merged_file)
    original_count = len(df)
    print(f"  Loaded {original_count} rows")

    # Remove empty runs (no hours and no drill distance)
    df_filtered, removed_empty_count = remove_empty_runs(df)
    after_empty_removal = len(df_filtered)

    # Detect duplicates (marks only Directional duplicates for removal)
    df_with_duplicates, rental_dup_count = detect_duplicates(df_filtered)

    # Get counts before removal
    directional_dup_count = df_with_duplicates['IS_DUPLICATE'].sum()

    # Remove Directional duplicates
    df_clean, removed_dup_count = remove_directional_duplicates(df_with_duplicates)
    final_count = len(df_clean)

    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"CLEAN_DD_MERGE_{timestamp}.xlsx"

    # Export to Excel (keeping IS_RENTAL_DUPLICATE temporarily for highlighting)
    print(f"\nExporting to: {output_file}")
    df_output = df_clean.drop(columns=['IS_DUPLICATE', 'SN_LAST_3'], errors='ignore')
    df_output.to_excel(output_file, index=False, engine='openpyxl')

    # Highlight Rental duplicates in yellow
    highlight_rental_duplicates(output_file, df_clean)

    # Remove IS_RENTAL_DUPLICATE column from final file
    print("\nRemoving temporary IS_RENTAL_DUPLICATE column from output...")
    wb = load_workbook(output_file)
    ws = wb.active

    # Find IS_RENTAL_DUPLICATE column
    rental_dup_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'IS_RENTAL_DUPLICATE':
            rental_dup_col = col
            break

    if rental_dup_col:
        ws.delete_cols(rental_dup_col)

    # Format DATE_IN and DATE_OUT as date-only (no time)
    print("Formatting DATE_IN and DATE_OUT columns (date only, no time)...")
    header_row = [cell.value for cell in ws[1]]
    date_in_col = None
    date_out_col = None

    for idx, header in enumerate(header_row, start=1):
        if header == 'DATE_IN':
            date_in_col = idx
        elif header == 'DATE_OUT':
            date_out_col = idx

    # Format DATE_IN column (date only)
    if date_in_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=date_in_col)
            if cell.value is not None:
                if hasattr(cell.value, 'date'):
                    cell.value = cell.value.date()
                cell.number_format = 'YYYY-MM-DD'

    # Format DATE_OUT column (date only)
    if date_out_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=date_out_col)
            if cell.value is not None:
                if hasattr(cell.value, 'date'):
                    cell.value = cell.value.date()
                cell.number_format = 'YYYY-MM-DD'

    print("  Applied date-only formatting to DATE_IN and DATE_OUT")

    wb.save(output_file)

    # Generate summary report
    generate_summary_report(original_count, after_empty_removal, directional_dup_count, rental_dup_count, final_count, output_file)

    print("\nCLEAN_DD_MERGE file created successfully!")
    print("Directional duplicates have been removed from the output.")
    print(f"Rental duplicates ({rental_dup_count}) are highlighted in YELLOW.")


if __name__ == "__main__":
    main()
