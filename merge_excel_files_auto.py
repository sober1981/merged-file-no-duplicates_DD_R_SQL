"""
Excel Files Merger Script (Auto-detect version)
This script automatically finds and merges Excel files based on filename patterns.
It searches for files starting with:
- Motor KPI
- CAM Run Tracker
- POG CAM
- POG MM

Data Corrections Applied:
1. JOB_TYPE: Motor KPI=Directional, CAM=Rental, POG blanks=Rental, MWD→Rental
2. DDS: Motor KPI=SDT, POG=Other, CAM=Keep as is
3. BHA: Blanks set to 1 (all sources)
4. RUN_NUM: Blanks set to 1 (all sources)
5. MY (CAM Run Tracker only):
   - Primary source: "Yield >45 Deg" (Column AP)
   - Fallback: "Yield 0-45 Deg" (Column AO) if AP blank
   - Text parsing: "18s" → 18.0, "11s to 15s" → 13.0 (average)

Author: Created for drilling optimization project
Date: 2025-10-30
Version: 2.3 (Auto-detect + Data Corrections + MY Column Enhancement)
"""

import pandas as pd
import numpy as np
from datetime import datetime
import warnings
import os
import glob
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

# File search patterns (case-insensitive)
FILE_PATTERNS = {
    'Motor_KPI': 'Motor KPI*.xlsx',
    'CAM_Run_Tracker': 'CAM Run Tracker*.xlsx',
    'POG_CAM_Usage': 'POG CAM*.xlsx',
    'POG_MM_Usage': 'POG MM*.xlsx'
}

MAPPING_FILE = 'FORMAT GRAL TABLE.xlsx'
BASIN_LOOKUP_FILE = 'LISTS_BASIN AND FORM_FAM.xlsx'
OUTPUT_FILE = f'MERGED_DATA_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

# ============================================================================
# AUTO-DETECT FILES
# ============================================================================

def find_files():
    """Automatically find required files based on patterns"""
    found_files = {}

    print("Searching for required files...")
    for key, pattern in FILE_PATTERNS.items():
        # Search for files matching pattern
        matches = glob.glob(pattern)

        if len(matches) == 0:
            print(f"  ERROR: No file found matching pattern '{pattern}'")
            return None
        elif len(matches) > 1:
            print(f"  WARNING: Multiple files found for '{pattern}':")
            for i, match in enumerate(matches, 1):
                print(f"    {i}. {match}")
            print(f"  Using: {matches[0]}")
            found_files[key] = matches[0]
        else:
            print(f"  Found: {matches[0]}")
            found_files[key] = matches[0]

    # Check for mapping and lookup files
    if not os.path.exists(MAPPING_FILE):
        print(f"  ERROR: Mapping file '{MAPPING_FILE}' not found!")
        return None
    else:
        print(f"  Found: {MAPPING_FILE}")

    if not os.path.exists(BASIN_LOOKUP_FILE):
        print(f"  ERROR: Lookup file '{BASIN_LOOKUP_FILE}' not found!")
        return None
    else:
        print(f"  Found: {BASIN_LOOKUP_FILE}")

    print("\nAll required files found successfully!\n")
    return found_files

# ============================================================================
# OPERATOR MAPPING
# ============================================================================

# Operator name standardization mapping (CAM Run Tracker -> Standard names)
OPERATOR_MAPPING = {
    'Aethon Energy': 'Aethon Energy Operating, LLC',
    'BPX': 'BPX Operating Company',
    'COMSTOCK RESOURCES': 'Comstock Oil & Gas LLP',
    'Camino': 'Camino Resources',
    'Caturus Energy': 'CATURUS ENERGY, LLC',
    'Comstock': 'Comstock Oil & Gas LLP',
    'Comstock Resources': 'Comstock Oil & Gas LLP',
    'Conoco': 'Conoco Phillips',
    'ConocoPhillips': 'Conoco Phillips',
    'Coterra': 'COTERRA',
    'Devon': 'Devon Energy',
    'Discovery': 'DISCOVERY NATURAL RESOURCES',
    'Exxon': 'EXXON',
    'Fervo': 'FERVO ENERGY COMPANY',
    'Greenlake Energy': 'GREENLAKE ENERGY',
    'Logos Operating LLC': 'LOGOS OPERATING LLC',
    'Mewbourne': 'Mewbourne Oil Company',
    'Mitsui': 'MITSUI E&P USA LLC',
    'Ovintiv': 'Ovintiv USA',
    'Oxy': 'OXY USA',
    'Oxy EOR': 'OXY USA',
    'Petro-Hunt': 'PETRO-HUNT',
    'Summit': 'Summit Petroleum',
    'XTO': 'EXXON'
}

# ============================================================================
# STEP 1: Load Mapping
# ============================================================================

def load_mapping():
    """Load the mapping configuration from FORMAT GRAL TABLE"""
    print("Loading mapping configuration...")
    df_mapping = pd.read_excel(MAPPING_FILE, sheet_name='Sheet1')

    # Create mapping dictionaries for each source
    mappings = {}
    target_headers = list(df_mapping.columns[1:])  # All columns except SOURCE

    for _, row in df_mapping.iterrows():
        source = row['SOURCE']
        mapping = {}
        for target_col in target_headers:
            source_col = row[target_col]
            if pd.notna(source_col) and str(source_col).strip() != '' and str(source_col) != 'Not Present':
                # Handle special case where mapping includes instructions
                if '(per' not in str(source_col):
                    mapping[source_col] = target_col
        mappings[source] = mapping
        print(f"  Loaded mappings for {source}")
        print(f"    {source}: {len(mapping)} mapped columns")

    return mappings, target_headers

# ============================================================================
# STEP 2: Load Lookup Tables
# ============================================================================

def load_lookup_tables():
    """Load basin and formation family lookup tables"""
    print("\nLoading lookup tables...")

    # Load Basin lookup
    basin_df = pd.read_excel(BASIN_LOOKUP_FILE, sheet_name='Basin')

    # Create a dictionary mapping county to basin
    # The Basin sheet has basin names as columns and counties as values
    county_to_basin = {}
    for col in basin_df.columns:
        basin_name = col
        for county in basin_df[col].dropna():
            county_to_basin[str(county).strip().upper()] = basin_name

    # Load Formation Family lookup
    df_formfam = pd.read_excel(BASIN_LOOKUP_FILE, sheet_name='FORM_FAM')

    print(f"  Loaded {len(county_to_basin)} county-to-basin mappings")
    print(f"  Loaded {len(df_formfam)} formation family mappings")

    return county_to_basin, df_formfam

# ============================================================================
# STEP 3: Read Source Files
# ============================================================================

def read_motor_kpi(file_path, mapping):
    """Read Motor KPI file"""
    print(f"\nReading Motor KPI file...")
    df = pd.read_excel(file_path)
    print(f"  Rows before cleaning: {len(df)}, Columns: {len(df.columns)}")

    # Check if headers are in the first row (like POG files)
    # If most columns are "Unnamed", headers are likely in first row
    unnamed_count = sum(1 for col in df.columns if str(col).startswith('Unnamed'))
    if unnamed_count > len(df.columns) * 0.5:  # More than 50% unnamed
        print(f"  Detected headers in first row, restructuring...")
        # The first row contains the actual headers
        new_headers = df.iloc[0]
        df = df[1:].copy()
        df.columns = new_headers

        # Remove any completely empty rows
        df = df.dropna(how='all')
        print(f"  Rows after cleaning: {len(df)}, Columns: {len(df.columns)}")

    # Save original BHA column before renaming
    original_bha = df['BHA'].copy() if 'BHA' in df.columns else None

    # Rename columns according to mapping
    df_renamed = df.rename(columns=mapping)
    df_renamed['SOURCE'] = 'Motor_KPI'

    # Restore BHA column
    if original_bha is not None:
        df_renamed['BHA'] = original_bha

    # Special handling: Map DATEIN and DATEOUT to DATE_IN and DATE_OUT
    # TIME_IN and TIME_OUT should come from the original TIME_IN/TIME_OUT columns (if they exist)
    # not from DATEIN/DATEOUT which are date-only fields
    if 'DATEIN' in df.columns:
        datein_dt = pd.to_datetime(df['DATEIN'], errors='coerce')
        df_renamed['DATE_IN'] = datein_dt.dt.date
    if 'DATEOUT' in df.columns:
        dateout_dt = pd.to_datetime(df['DATEOUT'], errors='coerce')
        df_renamed['DATE_OUT'] = dateout_dt.dt.date

    # TIME_IN and TIME_OUT are already mapped by the column mapping
    # But ensure they are preserved if they exist in the source
    if 'TIME_IN' in df.columns and ('TIME_IN' not in df_renamed.columns or df_renamed['TIME_IN'].isna().all()):
        df_renamed['TIME_IN'] = df['TIME_IN']
    if 'TIME_OUT' in df.columns and ('TIME_OUT' not in df_renamed.columns or df_renamed['TIME_OUT'].isna().all()):
        df_renamed['TIME_OUT'] = df['TIME_OUT']

    # Special handling: Map BENDANGLE to BEND and BEND_HSG
    if 'BENDANGLE' in df.columns:
        if 'BEND' not in df_renamed.columns or df_renamed['BEND'].isna().all():
            df_renamed['BEND'] = df['BENDANGLE']
        if 'BEND_HSG' not in df_renamed.columns or df_renamed['BEND_HSG'].isna().all():
            df_renamed['BEND_HSG'] = df['BENDANGLE']

    return df_renamed

def read_cam_run_tracker(file_path, mapping):
    """Read CAM Run Tracker file"""
    print(f"\nReading CAM Run Tracker file...")
    df = pd.read_excel(file_path, sheet_name='General')
    print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")

    # DEBUG: Print column names to check for Yield columns
    yield_cols = [col for col in df.columns if 'yield' in str(col).lower()]
    if yield_cols:
        print(f"  DEBUG: Found Yield columns: {yield_cols}")
    else:
        print(f"  DEBUG: No Yield columns found in file")

    # BEFORE renaming: Populate MY column with fallback logic
    # Primary: Column with "Yield" and ">45" (handles case variations and spaces)
    # Fallback: Column with "Yield" and "0-45" if primary is blank
    # Must do this BEFORE column renaming to access original column names

    # Find columns by pattern (case-insensitive, handles extra spaces)
    yield_high_col = None
    yield_low_col = None
    for col in df.columns:
        col_clean = str(col).strip().lower()
        if 'yield' in col_clean and '>45' in col_clean:
            yield_high_col = col  # Store original column name
        elif 'yield' in col_clean and '0-45' in col_clean:
            yield_low_col = col  # Store original column name

    if yield_high_col and yield_low_col:
        print(f"  Found: Primary='{yield_high_col}', Fallback='{yield_low_col}'")

        def get_my_value(row_orig):
            # Try primary source first (Yield >45)
            yield_high = row_orig.get(yield_high_col)
            if pd.notna(yield_high) and str(yield_high).strip() != '':
                return yield_high

            # Fallback to secondary source (Yield 0-45)
            yield_low = row_orig.get(yield_low_col)
            if pd.notna(yield_low) and str(yield_low).strip() != '':
                return yield_low

            return None

        # Create MY column in original df before renaming
        df['MY'] = df.apply(get_my_value, axis=1)

        # Count how many values came from each source for debugging
        total_my = df['MY'].notna().sum()
        high_populated = df[yield_high_col].notna() & (df[yield_high_col].astype(str).str.strip() != '')
        low_populated = df[yield_low_col].notna() & (df[yield_low_col].astype(str).str.strip() != '')
        from_high = (high_populated & df['MY'].notna()).sum()
        from_low = (~high_populated & low_populated & df['MY'].notna()).sum()

        print(f"  Populated MY column: {total_my} total values ({from_high} from primary, {from_low} from fallback)")

        # Remove any mapping entries that would create duplicate MY column
        # Since we already created MY with fallback logic, prevent mapping from creating it again
        keys_to_remove = [k for k, v in mapping.items() if v == 'MY']
        for key in keys_to_remove:
            print(f"  Removing mapping '{key}' -> 'MY' to prevent duplicate column")
            del mapping[key]
    else:
        print(f"  WARNING: Could not find both Yield columns (High={yield_high_col}, Low={yield_low_col})")

    # Rename columns according to mapping
    df_renamed = df.rename(columns=mapping)
    df_renamed['SOURCE'] = 'CAM_Run_Tracker'

    # Special handling: Map Run # to BHA
    if 'Run #' in df.columns:
        if 'BHA' not in df_renamed.columns or df_renamed['BHA'].isna().all():
            df_renamed['BHA'] = df['Run #']

    # Special handling: Split Start of Run and End of Run into DATE and TIME
    if 'Start of Run' in df.columns:
        start_datetime = pd.to_datetime(df['Start of Run'], errors='coerce')
        df_renamed['DATE_IN'] = start_datetime.dt.date
        df_renamed['TIME_IN'] = start_datetime.dt.time

    if 'End of Run' in df.columns:
        end_datetime = pd.to_datetime(df['End of Run'], errors='coerce')
        df_renamed['DATE_OUT'] = end_datetime.dt.date
        df_renamed['TIME_OUT'] = end_datetime.dt.time

    # Special handling: Map Bend to BEND and BEND_HSG
    if 'Bend' in df.columns:
        if 'BEND' not in df_renamed.columns or df_renamed['BEND'].isna().all():
            df_renamed['BEND'] = df['Bend']
        if 'BEND_HSG' not in df_renamed.columns or df_renamed['BEND_HSG'].isna().all():
            df_renamed['BEND_HSG'] = df['Bend']

    return df_renamed

def read_pog_cam_usage(file_path, mapping):
    """Read POG CAM Usage file"""
    print(f"\nReading POG CAM Usage file...")
    df = pd.read_excel(file_path, sheet_name='POG Tool Usage')
    print(f"  Rows before cleaning: {len(df)}")

    # The first row contains the actual headers
    new_headers = df.iloc[0]
    df = df[1:].copy()
    df.columns = new_headers

    # Remove any completely empty rows
    df = df.dropna(how='all')

    print(f"  Rows after cleaning: {len(df)}, Columns: {len(df.columns)}")

    # Rename columns according to mapping
    df_renamed = df.rename(columns=mapping)
    df_renamed['SOURCE'] = 'POG_CAM_Usage'

    # Special handling: Map Brt Date and Art Date to DATE_IN and DATE_OUT
    if 'Brt Date' in df.columns:
        if 'DATE_IN' not in df_renamed.columns or df_renamed['DATE_IN'].isna().all():
            df_renamed['DATE_IN'] = pd.to_datetime(df['Brt Date'], errors='coerce').dt.date
    if 'Art Date' in df.columns:
        if 'DATE_OUT' not in df_renamed.columns or df_renamed['DATE_OUT'].isna().all():
            df_renamed['DATE_OUT'] = pd.to_datetime(df['Art Date'], errors='coerce').dt.date

    # Special handling: Map Fixed or Adjustable to BEND (use whichever has a value)
    if 'Fixed' in df.columns or 'Adjustable' in df.columns:
        if 'BEND' not in df_renamed.columns or df_renamed['BEND'].isna().all():
            # Try Fixed first, then Adjustable as fallback
            if 'Fixed' in df.columns:
                df_renamed['BEND'] = df['Fixed']
            if 'Adjustable' in df.columns:
                # Fill BEND with Adjustable where Fixed is empty
                df_renamed['BEND'] = df_renamed['BEND'].fillna(df['Adjustable'])

        # Also populate BEND_HSG with the same values
        if 'BEND_HSG' not in df_renamed.columns or df_renamed['BEND_HSG'].isna().all():
            df_renamed['BEND_HSG'] = df_renamed['BEND']

    # Special handling: Map Job Type column to JOB_TYPE
    if 'Job Type' in df.columns:
        if 'JOB_TYPE' not in df_renamed.columns or df_renamed['JOB_TYPE'].isna().all():
            df_renamed['JOB_TYPE'] = df['Job Type']

    return df_renamed

def read_pog_mm_usage(file_path, mapping):
    """Read POG MM Usage file"""
    print(f"\nReading POG MM Usage file...")
    df = pd.read_excel(file_path, sheet_name='POG Tool Usage')
    print(f"  Rows before cleaning: {len(df)}")

    # The first row contains the actual headers
    new_headers = df.iloc[0]
    df = df[1:].copy()
    df.columns = new_headers

    # Remove any completely empty rows
    df = df.dropna(how='all')

    print(f"  Rows after cleaning: {len(df)}, Columns: {len(df.columns)}")

    # Rename columns according to mapping
    df_renamed = df.rename(columns=mapping)
    df_renamed['SOURCE'] = 'POG_MM_Usage'

    # Special handling: Map Brt Date and Art Date to DATE_IN and DATE_OUT
    if 'Brt Date' in df.columns:
        if 'DATE_IN' not in df_renamed.columns or df_renamed['DATE_IN'].isna().all():
            df_renamed['DATE_IN'] = pd.to_datetime(df['Brt Date'], errors='coerce').dt.date
    if 'Art Date' in df.columns:
        if 'DATE_OUT' not in df_renamed.columns or df_renamed['DATE_OUT'].isna().all():
            df_renamed['DATE_OUT'] = pd.to_datetime(df['Art Date'], errors='coerce').dt.date

    # Special handling: Map Fixed or Adjustable to BEND (use whichever has a value)
    if 'Fixed' in df.columns or 'Adjustable' in df.columns:
        if 'BEND' not in df_renamed.columns or df_renamed['BEND'].isna().all():
            # Try Fixed first, then Adjustable as fallback
            if 'Fixed' in df.columns:
                df_renamed['BEND'] = df['Fixed']
            if 'Adjustable' in df.columns:
                # Fill BEND with Adjustable where Fixed is empty
                df_renamed['BEND'] = df_renamed['BEND'].fillna(df['Adjustable'])

        # Also populate BEND_HSG with the same values
        if 'BEND_HSG' not in df_renamed.columns or df_renamed['BEND_HSG'].isna().all():
            df_renamed['BEND_HSG'] = df_renamed['BEND']

    # Special handling: Map Job Type column to JOB_TYPE
    if 'Job Type' in df.columns:
        if 'JOB_TYPE' not in df_renamed.columns or df_renamed['JOB_TYPE'].isna().all():
            df_renamed['JOB_TYPE'] = df['Job Type']

    return df_renamed

# ============================================================================
# STEP 4: Clean County Names
# ============================================================================

def clean_county_names(df, source_name):
    """Extract STATE from county and clean county names"""
    if 'COUNTY' not in df.columns:
        return df

    # Only apply to Motor KPI and POG files
    if source_name in ['Motor_KPI', 'POG_CAM_Usage', 'POG_MM_Usage']:
        print(f"  Extracting STATE from county and cleaning county names...")

        import re

        def extract_state_and_clean_county(county_str):
            if pd.isna(county_str):
                return None, None

            county_str = str(county_str).strip()

            # Extract state (last 2 capital letters)
            state_match = re.search(r'\b([A-Z]{2})$', county_str)
            state = state_match.group(1) if state_match else None

            # Clean county name
            # Remove "County", "Parish", and state abbreviation
            cleaned = county_str
            cleaned = re.sub(r'\s+County\s*', ' ', cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r'\s+Parish\s*', ' ', cleaned, flags=re.IGNORECASE)
            if state:
                cleaned = re.sub(r'\s+' + state + r'\s*$', '', cleaned)

            return state, cleaned.strip()

        # Extract states and clean counties
        states = []
        cleaned_counties = []

        for county in df['COUNTY']:
            state, cleaned = extract_state_and_clean_county(county)
            states.append(state)
            cleaned_counties.append(cleaned)

        # Update STATE column
        if 'STATE' in df.columns:
            df['STATE'] = df['STATE'].fillna(pd.Series(states))
        else:
            df['STATE'] = states

        # Update COUNTY column
        df['COUNTY'] = cleaned_counties

        print(f"    Extracted STATE for {sum(1 for s in states if s is not None)} records")
        print(f"    Cleaned {len(cleaned_counties)} county names")

        # Print sample
        print(f"    Sample results:")
        for i in range(min(5, len(df))):
            if pd.notna(df.iloc[i]['COUNTY']):
                print(f"      County: {df.iloc[i]['COUNTY']}, State: {df.iloc[i]['STATE']}")

    return df

# ============================================================================
# STEP 5: Standardize Operator Names
# ============================================================================

def standardize_operator_names(df, source_name):
    """Standardize operator names for CAM Run Tracker data"""
    if 'OPERATOR' not in df.columns:
        return df

    # Only apply mapping to CAM Run Tracker data
    if source_name == 'CAM_Run_Tracker':
        print(f"  Standardizing operator names...")

        # Count changes before
        original_names = df['OPERATOR'].value_counts()
        changes_made = 0

        # Apply mapping
        for old_name, new_name in OPERATOR_MAPPING.items():
            mask = df['OPERATOR'] == old_name
            count = mask.sum()
            if count > 0:
                df.loc[mask, 'OPERATOR'] = new_name
                changes_made += 1
                print(f"    {old_name} -> {new_name} ({count} records)")

        print(f"  Total operator names standardized: {changes_made}")

    return df

# ============================================================================
# STEP 6: Format Dates and Create START_DATE/END_DATE
# ============================================================================

def format_dates_and_datetimes(df):
    """Format dates and create START_DATE/END_DATE columns"""

    # Format DATE_IN and DATE_OUT to proper date format
    if 'DATE_IN' in df.columns:
        df['DATE_IN'] = pd.to_datetime(df['DATE_IN'], errors='coerce').dt.date
    if 'DATE_OUT' in df.columns:
        df['DATE_OUT'] = pd.to_datetime(df['DATE_OUT'], errors='coerce').dt.date

    print(f"  Formatted DATE_IN/DATE_OUT to date format")

    # Create START_DATE and END_DATE
    if 'START_DATE' in df.columns and 'END_DATE' in df.columns:
        def create_start_date(row):
            # For Motor KPI: combine DATE_IN + TIME_IN
            if row['SOURCE'] in ['Motor_KPI']:
                if pd.notna(row['DATE_IN']) and pd.notna(row.get('TIME_IN')):
                    try:
                        # TIME_IN might be a string like '09:00:00', convert to time object
                        time_in = row['TIME_IN']
                        if isinstance(time_in, str):
                            # Parse string time to time object
                            time_in = pd.to_datetime(time_in, format='%H:%M:%S').time()

                        # Combine date and time
                        return pd.Timestamp.combine(row['DATE_IN'], time_in)
                    except Exception as e:
                        # If combination fails, just use the date
                        return pd.to_datetime(row['DATE_IN'])
            # For POG files without time
            elif row['SOURCE'] in ['POG_CAM_Usage', 'POG_MM_Usage']:
                if pd.notna(row['DATE_IN']):
                    return pd.to_datetime(row['DATE_IN'])
            # For CAM Run Tracker, START_DATE already populated
            return row['START_DATE']

        def create_end_date(row):
            # For Motor KPI: combine DATE_OUT + TIME_OUT
            if row['SOURCE'] in ['Motor_KPI']:
                if pd.notna(row['DATE_OUT']) and pd.notna(row.get('TIME_OUT')):
                    try:
                        # TIME_OUT might be a string like '10:00:00', convert to time object
                        time_out = row['TIME_OUT']
                        if isinstance(time_out, str):
                            # Parse string time to time object
                            time_out = pd.to_datetime(time_out, format='%H:%M:%S').time()

                        # Combine date and time
                        return pd.Timestamp.combine(row['DATE_OUT'], time_out)
                    except Exception as e:
                        # If combination fails, just use the date
                        return pd.to_datetime(row['DATE_OUT'])
            # For POG files without time
            elif row['SOURCE'] in ['POG_CAM_Usage', 'POG_MM_Usage']:
                if pd.notna(row['DATE_OUT']):
                    return pd.to_datetime(row['DATE_OUT'])
            # For CAM Run Tracker, END_DATE already populated
            return row['END_DATE']

        df['START_DATE'] = df.apply(create_start_date, axis=1)
        df['END_DATE'] = df.apply(create_end_date, axis=1)

        print(f"  Created START_DATE/END_DATE from date+time combinations")

    return df

# ============================================================================
# STEP 7: Populate LOBE/STAGE and DDS
# ============================================================================

def populate_lobe_stage_and_dds(df):
    """
    Populate LOBE/STAGE column by combining LOBES and STAGES
    Populate DDS column based on source-specific logic
    """

    # Handle LOBE/STAGE column
    if 'LOBE/STAGE' in df.columns and 'LOBES' in df.columns and 'STAGES' in df.columns:
        def combine_lobe_stage(row):
            # For Motor KPI and POG files, combine LOBES + ":" + STAGES
            # LOBES should be like "6/7" and STAGES should be like "7.8"
            # Result: "6/7:7.8"
            if row['SOURCE'] in ['Motor_KPI', 'POG_CAM_Usage', 'POG_MM_Usage']:
                lobe = row['LOBES']
                stage = row['STAGES']
                if pd.notna(lobe) and pd.notna(stage):
                    # Convert to string and ensure proper format
                    lobe_str = str(lobe).strip()
                    stage_str = str(stage).strip()

                    # If LOBES already contains ":", it might be formatted wrong
                    # We want LOBE/STAGE format (e.g., "6/7"), not LOBES:STAGES
                    # Remove any existing ":" from lobe if present
                    if ':' in lobe_str:
                        # This means LOBES was mistakenly formatted as "6:7" instead of "6/7"
                        lobe_str = lobe_str.replace(':', '/')

                    return f"{lobe_str}:{stage_str}"
            # CAM Run Tracker: replace "-" with ":" to match format
            elif row['SOURCE'] == 'CAM_Run_Tracker':
                if pd.notna(row['LOBE/STAGE']):
                    return str(row['LOBE/STAGE']).replace('-', ':')
            return row['LOBE/STAGE']

        df['LOBE/STAGE'] = df.apply(combine_lobe_stage, axis=1)
        print(f"  Combined LOBES and STAGES into LOBE/STAGE column")

    # Handle DDS column
    if 'DDS' in df.columns:
        def populate_dds(row):
            source = row['SOURCE']

            # Motor KPI: Always "SDT"
            if source == 'Motor_KPI':
                return 'SDT'

            # CAM Run Tracker: Keep as is (extract first complete word/company name)
            elif source == 'CAM_Run_Tracker':
                if pd.notna(row['DDS']):
                    dds_value = str(row['DDS']).strip()
                    # Extract first word before space or /
                    import re
                    match = re.match(r'^([A-Za-z]+)', dds_value)
                    if match:
                        return match.group(1)
                return row['DDS']

            # POG files: Always "Other"
            elif source in ['POG_CAM_Usage', 'POG_MM_Usage']:
                return 'Other'

            return row['DDS'] if pd.notna(row['DDS']) else None

        df['DDS'] = df.apply(populate_dds, axis=1)
        print(f"  Populated DDS column (Motor KPI=SDT, POG=Other, CAM=Keep as is)")

    return df

# ============================================================================
# STEP 10: Populate TOTAL HRS
# ============================================================================

def populate_total_hrs(df):
    """
    Populate Total Hrs (C+D) column:
    - Motor KPI: CIRC_HOURS + DRILLING_HOURS
    - CAM Run Tracker: Already populated
    - POG files: Already populated
    """
    total_hrs_col = 'Total Hrs (C+D)'
    if total_hrs_col in df.columns:
        def calculate_total_hrs(row):
            # For Motor KPI, sum CIRC_HOURS and DRILLING_HOURS
            if row['SOURCE'] == 'Motor_KPI':
                circ = row.get('CIRC_HOURS', 0) if pd.notna(row.get('CIRC_HOURS')) else 0
                drilling = row.get('DRILLING_HOURS', 0) if pd.notna(row.get('DRILLING_HOURS')) else 0
                return circ + drilling
            # For other sources, keep existing value
            return row[total_hrs_col]

        df[total_hrs_col] = df.apply(calculate_total_hrs, axis=1)
        print(f"  Calculated Total Hrs (C+D) from CIRC_HOURS + DRILLING_HOURS for Motor KPI")

    return df

# ============================================================================
# STEP 11: Add UPDATE Column
# ============================================================================

def add_update_column(df):
    """
    Add UPDATE column with today's date (date when merge is performed)
    """
    from datetime import datetime

    if 'UPDATE' in df.columns:
        df['UPDATE'] = datetime.now().date()
        print(f"  Added UPDATE column with merge date: {datetime.now().date()}")

    return df

# ============================================================================
# STEP 12: Populate MOTOR_TYPE2
# ============================================================================

def populate_motor_type2(df):
    """
    Populate MOTOR_TYPE2 column based on source-specific logic:

    Motor KPI:
    - If "MLA07" in SN -> "CAM DD"
    - If "TDI" in MOTOR_MAKE and no "MLA07" in SN -> "TDI CONV"
    - If no "TDI" in MOTOR_MAKE -> "3RD PARTY"

    CAM Run Tracker:
    - All -> "CAM RENTAL"

    POG_CAM:
    - If JOB_TYPE is "RENTAL" -> "CAM RENTAL"
    - If JOB_TYPE is "DIRECTIONAL" -> "CAM DD"

    POG_MM:
    - All -> "TDI CONV"
    """
    if 'MOTOR_TYPE2' in df.columns:
        def determine_motor_type2(row):
            source = row['SOURCE']

            # Motor KPI logic
            if source == 'Motor_KPI':
                sn = str(row.get('SN', '')).upper() if pd.notna(row.get('SN')) else ''
                motor_make = str(row.get('MOTOR_MAKE', '')).upper() if pd.notna(row.get('MOTOR_MAKE')) else ''

                if 'MLA07' in sn:
                    return 'CAM DD'
                elif 'TDI' in motor_make and 'MLA07' not in sn:
                    return 'TDI CONV'
                else:
                    return '3RD PARTY'

            # CAM Run Tracker logic
            elif source == 'CAM_Run_Tracker':
                return 'CAM RENTAL'

            # POG_CAM logic
            elif source == 'POG_CAM_Usage':
                job_type = str(row.get('JOB_TYPE', '')).strip().upper() if pd.notna(row.get('JOB_TYPE')) else ''
                if 'RENTAL' in job_type:
                    return 'CAM RENTAL'
                elif 'DIRECTIONAL' in job_type:
                    return 'CAM DD'
                return None

            # POG_MM logic
            elif source == 'POG_MM_Usage':
                return 'TDI CONV'

            return None

        df['MOTOR_TYPE2'] = df.apply(determine_motor_type2, axis=1)
        print(f"  Populated MOTOR_TYPE2 column based on source-specific logic")

    return df

# ============================================================================
# STEP 13: Populate and Clean JOB_TYPE
# ============================================================================

def populate_and_clean_job_type(df):
    """
    Populate and standardize JOB_TYPE for all sources:
    - Motor KPI: All should be "Directional"
    - CAM Run Tracker: All should be "Rental"
    - POG files: If blank, set to "Rental"
    - Replace "MWD" with "Rental"
    - Replace "Directional- MWD and Motor" with "Directional"
    - Only two allowed values: "Directional" or "Rental"
    """
    if 'JOB_TYPE' in df.columns:
        def clean_job_type(row):
            source = row['SOURCE']
            job_type = row.get('JOB_TYPE', None)

            # Motor KPI: All are Directional
            if source == 'Motor_KPI':
                return 'Directional'

            # CAM Run Tracker: All are Rental
            if source == 'CAM_Run_Tracker':
                return 'Rental'

            # For POG files and other sources, clean the existing value
            if pd.notna(job_type):
                job_type_str = str(job_type).strip()

                # Replace "MWD" with "Rental"
                if job_type_str.upper() == 'MWD':
                    return 'Rental'

                # Replace "Directional- MWD and Motor" with "Directional"
                if 'Directional- MWD and Motor' in job_type_str:
                    return 'Directional'

                # Standardize to exact case
                if job_type_str.upper() == 'DIRECTIONAL':
                    return 'Directional'
                if job_type_str.upper() == 'RENTAL':
                    return 'Rental'

                return job_type_str

            # POG files: If blank, set to "Rental"
            if source in ['POG_CAM_Usage', 'POG_MM_Usage']:
                return 'Rental'

            return job_type

        df['JOB_TYPE'] = df.apply(clean_job_type, axis=1)
        print(f"  Populated and standardized JOB_TYPE (Motor KPI=Directional, CAM=Rental, POG blanks=Rental)")

    return df

# ============================================================================
# STEP 14: Populate MOTOR_MODEL
# ============================================================================

def populate_motor_model(df):
    """
    Populate MOTOR_MODEL column based on source-specific logic:

    Motor KPI:
    - If MOTOR_MAKE is TDI: Extract model from SN (475, 500, 650, 712, 800, 962)
    - If MOTOR_MAKE is not TDI: Use MOTOR_OD

    CAM Run Tracker:
    - Keep existing MOTOR_MODEL

    POG files:
    - Convert text to numbers:
      5 → 500
      5-4/4 → 575
      6-1/2 → 650
      7-1/8 → 712
      8 → 800
      9-5/8 → 962
    """
    if 'MOTOR_MODEL' in df.columns:
        def calculate_motor_model(row):
            source = row['SOURCE']

            # Motor KPI logic
            if source == 'Motor_KPI':
                motor_make = str(row.get('MOTOR_MAKE', '')).upper() if pd.notna(row.get('MOTOR_MAKE')) else ''

                if 'TDI' in motor_make:
                    # Extract model from SN (serial number)
                    sn = str(row.get('SN', '')).upper() if pd.notna(row.get('SN')) else ''

                    # Look for 3-digit numbers in SN that match motor models
                    import re
                    # Common TDI models: 475, 500, 575, 650, 712, 800, 962
                    matches = re.findall(r'\b(475|500|575|650|712|800|962)\b', sn)
                    if matches:
                        return matches[0]  # Return first match

                    # If no match found, return existing MOTOR_MODEL or None
                    return row.get('MOTOR_MODEL')
                else:
                    # Not TDI, use MOTOR_OD
                    motor_od = row.get('MOTOR_OD')
                    if pd.notna(motor_od):
                        return str(motor_od).strip()
                    return row.get('MOTOR_MODEL')

            # CAM Run Tracker: keep existing
            elif source == 'CAM_Run_Tracker':
                return row.get('MOTOR_MODEL')

            # POG files: convert text to numbers
            elif source in ['POG_CAM_Usage', 'POG_MM_Usage']:
                motor_model = row.get('MOTOR_MODEL')
                if pd.notna(motor_model):
                    model_str = str(motor_model).strip()

                    # Conversion mapping
                    conversions = {
                        '5': '500',
                        '5-4/4': '575',
                        '6-1/2': '650',
                        '7-1/8': '712',
                        '8': '800',
                        '9-5/8': '962'
                    }

                    # Check for exact match
                    if model_str in conversions:
                        return conversions[model_str]

                    # Check if already a number
                    if model_str.isdigit():
                        return model_str

                    return model_str
                return None

            return row.get('MOTOR_MODEL')

        df['MOTOR_MODEL'] = df.apply(calculate_motor_model, axis=1)
        print(f"  Populated MOTOR_MODEL based on source-specific logic")

    return df

# ============================================================================
# STEP 15: Parse MY Column Text for CAM Run Tracker
# ============================================================================

def parse_my_column_text(df):
    """
    Parse MY column text patterns for CAM Run Tracker rows only.

    Patterns handled:
    - "18s" or "18S" → 18.0
    - "11s to 15s" → 13.0 (average of 11 and 15)
    - Case-insensitive
    - Unknown patterns → Leave as is

    Result must be numeric in merged file.
    """
    import re

    if 'MY' not in df.columns:
        return df

    print("\nParsing MY column text for CAM Run Tracker rows...")

    def parse_my_value(row):
        # Only process CAM Run Tracker rows
        if row['SOURCE'] != 'CAM_Run_Tracker':
            return row['MY']

        my_value = row['MY']

        # If already numeric or empty, return as is
        if pd.isna(my_value):
            return my_value

        # Convert to string for processing
        my_str = str(my_value).strip()

        # If it's already a number, return it
        try:
            return float(my_str)
        except ValueError:
            pass

        # Pattern 1: Single number with 's' or 'S' (e.g., "18s", "18S")
        # Match: number followed by optional 's' or 'S'
        pattern1 = r'^(\d+(?:\.\d+)?)[sS]?$'
        match1 = re.match(pattern1, my_str, re.IGNORECASE)
        if match1:
            return float(match1.group(1))

        # Pattern 2: Range with "to" (e.g., "11s to 15s", "11 to 15")
        # Match: number (optional s) + "to" + number (optional s)
        pattern2 = r'^(\d+(?:\.\d+)?)[sS]?\s+to\s+(\d+(?:\.\d+)?)[sS]?$'
        match2 = re.match(pattern2, my_str, re.IGNORECASE)
        if match2:
            num1 = float(match2.group(1))
            num2 = float(match2.group(2))
            return (num1 + num2) / 2.0

        # If no pattern matched, return original value (leave as is)
        return my_value

    df['MY'] = df.apply(parse_my_value, axis=1)
    print(f"  Parsed MY column text patterns for CAM Run Tracker rows")

    return df

# ============================================================================
# STEP 16: Populate BHA and RUN_NUM
# ============================================================================

def populate_bha_and_run_num(df):
    """
    Populate BHA and RUN_NUM columns with default values where blank:
    - POG files: If blank, set BHA and RUN_NUM to 1
    - Motor KPI & CAM Run Tracker: If blank, set BHA and RUN_NUM to 1
    - Keep existing values if present
    """

    # Handle BHA column
    if 'BHA' in df.columns:
        def populate_bha(row):
            source = row['SOURCE']
            bha = row.get('BHA', None)

            # If blank or NaN, set to 1
            if pd.isna(bha) or bha == '':
                return 1

            # Keep existing value
            return bha

        df['BHA'] = df.apply(populate_bha, axis=1)
        print(f"  Populated BHA column (blanks set to 1)")

    # Handle RUN_NUM column
    if 'RUN_NUM' in df.columns:
        def populate_run_num(row):
            source = row['SOURCE']
            run_num = row.get('RUN_NUM', None)

            # If blank or NaN, set to 1
            if pd.isna(run_num) or run_num == '':
                return 1

            # Keep existing value
            return run_num

        df['RUN_NUM'] = df.apply(populate_run_num, axis=1)
        print(f"  Populated RUN_NUM column (blanks set to 1)")

    return df

# ============================================================================
# STEP 16: Convert Numeric Columns to Text
# ============================================================================

def convert_to_text_format(df):
    """
    Convert specific numeric columns to text format for standardization.

    Columns to convert:
    - MOTOR_MODEL
    - BEND
    - BEND_HSG

    This ensures these values are stored as text, not numbers.
    """
    columns_to_convert = ['MOTOR_MODEL', 'BEND', 'BEND_HSG']

    for col in columns_to_convert:
        if col in df.columns:
            # Convert to string, handling NaN values
            df[col] = df[col].apply(lambda x: str(x) if pd.notna(x) else x)

    print(f"  Converted MOTOR_MODEL, BEND, BEND_HSG to text format")

    return df

# ============================================================================
# STEP 8: Apply Lookups
# ============================================================================

def apply_basin_lookup(df, county_to_basin):
    """Apply basin lookup based on county"""
    if 'COUNTY' in df.columns and 'BASIN' in df.columns:
        def get_basin(county):
            if pd.isna(county):
                return None
            county_str = str(county).strip().upper()
            return county_to_basin.get(county_str, None)

        df['BASIN'] = df['COUNTY'].apply(get_basin)

    return df

def apply_formfam_lookup(df, formfam_df):
    """Apply formation family lookup"""
    if 'FORMATION' in df.columns and 'BASIN' in df.columns and 'FORM_FAM' in df.columns:
        # Create lookup dictionary
        formfam_dict = {}
        for _, row in formfam_df.iterrows():
            key = (str(row['Basin']).upper(), str(row['Keyword']).upper())
            formfam_dict[key] = row['Formation Family']

        def get_form_fam(row):
            if pd.isna(row['FORMATION']) or pd.isna(row['BASIN']):
                return None

            basin = str(row['BASIN']).upper()
            formation = str(row['FORMATION']).upper()

            # Look for keyword match in formation name
            for (lookup_basin, keyword), form_fam in formfam_dict.items():
                if lookup_basin == basin and keyword in formation:
                    return form_fam

            return None

        df['FORM_FAM'] = df.apply(get_form_fam, axis=1)

    return df

# ============================================================================
# STEP 5: Merge All Data
# ============================================================================

def merge_all_files(FILES):
    """Main function to merge all files"""

    print("="*80)
    print("EXCEL FILES MERGER - STARTING")
    print("="*80)

    # Step 1: Load mapping
    mappings, target_headers = load_mapping()

    # Step 2: Load lookup tables
    county_to_basin, formfam_df = load_lookup_tables()

    # Step 3: Read all source files
    dfs = []

    # Motor KPI
    df_motor = read_motor_kpi(FILES['Motor_KPI'], mappings['Motor_KPI'])
    df_motor = clean_county_names(df_motor, 'Motor_KPI')
    df_motor = standardize_operator_names(df_motor, 'Motor_KPI')
    dfs.append(df_motor)

    # CAM Run Tracker
    df_cam = read_cam_run_tracker(FILES['CAM_Run_Tracker'], mappings['CAM Run Tracker'])
    df_cam = clean_county_names(df_cam, 'CAM_Run_Tracker')
    df_cam = standardize_operator_names(df_cam, 'CAM_Run_Tracker')
    dfs.append(df_cam)

    # POG CAM Usage
    df_pog_cam = read_pog_cam_usage(FILES['POG_CAM_Usage'], mappings['POG_CAM_Usage'])
    df_pog_cam = clean_county_names(df_pog_cam, 'POG_CAM_Usage')
    df_pog_cam = standardize_operator_names(df_pog_cam, 'POG_CAM_Usage')
    dfs.append(df_pog_cam)

    # POG MM Usage
    df_pog_mm = read_pog_mm_usage(FILES['POG_MM_Usage'], mappings['POG_MM_Usage'])
    df_pog_mm = clean_county_names(df_pog_mm, 'POG_MM_Usage')
    df_pog_mm = standardize_operator_names(df_pog_mm, 'POG_MM_Usage')
    dfs.append(df_pog_mm)

    # Step 4: Concatenate all dataframes
    print("\n" + "="*80)
    print("MERGING DATA")
    print("="*80)

    # NOTE: sort=False preserves the original row order from each source file
    # This ensures CAM Run Tracker rows maintain their original order
    df_merged = pd.concat(dfs, ignore_index=True, sort=False)
    print(f"\nTotal rows after merge: {len(df_merged)}")
    print(f"Total columns: {len(df_merged.columns)}")

    # Step 5: Ensure all target headers are present (add missing columns with NaN)
    for header in target_headers:
        if header not in df_merged.columns:
            df_merged[header] = np.nan

    # Step 6: Reorder columns to match target format (keep SOURCE column)
    # SOURCE is not in target_headers but we need it for transformations
    columns_to_keep = ['SOURCE'] + target_headers
    df_merged = df_merged[columns_to_keep]

    # Step 7: Apply lookups
    print("\nApplying lookup tables...")
    df_merged = apply_basin_lookup(df_merged, county_to_basin)
    df_merged = apply_formfam_lookup(df_merged, formfam_df)

    # Step 8: Format dates and create START_DATE/END_DATE
    print("\nFormatting dates and creating START_DATE/END_DATE...")
    df_merged = format_dates_and_datetimes(df_merged)

    # Step 9: Populate LOBE/STAGE and DDS columns
    print("\nPopulating LOBE/STAGE and DDS columns...")
    df_merged = populate_lobe_stage_and_dds(df_merged)

    # Step 10: Populate TOTAL HRS
    print("\nPopulating TOTAL HRS...")
    df_merged = populate_total_hrs(df_merged)

    # Step 11: Add UPDATE column
    print("\nAdding UPDATE column...")
    df_merged = add_update_column(df_merged)

    # Step 12: Populate MOTOR_TYPE2
    print("\nPopulating MOTOR_TYPE2...")
    df_merged = populate_motor_type2(df_merged)

    # Step 13: Populate and Clean JOB_TYPE
    print("\nPopulating and cleaning JOB_TYPE...")
    df_merged = populate_and_clean_job_type(df_merged)

    # Step 14: Populate MOTOR_MODEL
    print("\nPopulating MOTOR_MODEL...")
    df_merged = populate_motor_model(df_merged)

    # Step 15: Parse MY column text for CAM Run Tracker
    df_merged = parse_my_column_text(df_merged)

    # Step 16: Populate BHA and RUN_NUM
    print("\nPopulating BHA and RUN_NUM...")
    df_merged = populate_bha_and_run_num(df_merged)

    # Step 17: Convert numeric columns to text format
    print("\nConverting numeric columns to text format...")
    df_merged = convert_to_text_format(df_merged)

    # Step 18: Export to Excel
    print("\n" + "="*80)
    print("EXPORTING RESULTS")
    print("="*80)

    print(f"\nWriting to: {OUTPUT_FILE}")

    # Use ExcelWriter with openpyxl for better compatibility with Spotfire
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import numbers

    # Create workbook and write data
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
        # Write DataFrame to Excel
        df_merged.to_excel(writer, index=False, sheet_name='Merged Data')

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Merged Data']

        # Set properties for better compatibility
        workbook.properties.creator = "Scout Downhole - Drilling Optimization"
        workbook.properties.title = "Merged Scorecard Data"
        workbook.properties.description = "Merged drilling scorecard data from multiple sources"

        # Find DATE_IN and DATE_OUT column indices (K and L = columns 11 and 12)
        header_row = [cell.value for cell in worksheet[1]]
        date_in_col = None
        date_out_col = None
        start_date_col = None
        end_date_col = None

        for idx, header in enumerate(header_row, start=1):
            if header == 'DATE_IN':
                date_in_col = idx
            elif header == 'DATE_OUT':
                date_out_col = idx
            elif header == 'START_DATE':
                start_date_col = idx
            elif header == 'END_DATE':
                end_date_col = idx

        # Format DATE_IN and DATE_OUT as DATE ONLY (no time)
        # Also convert datetime values to date-only values
        if date_in_col:
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=date_in_col)
                if cell.value is not None:
                    # If it's a datetime, convert to date only
                    if hasattr(cell.value, 'date'):
                        cell.value = cell.value.date()
                    # Set the number format to display as date only
                    cell.number_format = 'YYYY-MM-DD'

        if date_out_col:
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=date_out_col)
                if cell.value is not None:
                    # If it's a datetime, convert to date only
                    if hasattr(cell.value, 'date'):
                        cell.value = cell.value.date()
                    # Set the number format to display as date only
                    cell.number_format = 'YYYY-MM-DD'

        # Format START_DATE and END_DATE as DATE + TIME
        if start_date_col:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=start_date_col)
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'  # Date and time format

        if end_date_col:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=end_date_col)
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'  # Date and time format

        print(f"  Applied date formatting: DATE_IN/OUT=date only, START/END_DATE=date+time")

        # Auto-adjust column widths for readability
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"  Excel file created with openpyxl engine and optimized formatting")

    print("\n" + "="*80)
    print("MERGE COMPLETE!")
    print("="*80)
    print(f"\nOutput file: {OUTPUT_FILE}")
    print(f"Total rows: {len(df_merged)}")
    print(f"Total columns: {len(df_merged.columns)}")

    # Print summary by source
    print("\n" + "-"*80)
    print("DATA SUMMARY BY SOURCE")
    print("-"*80)
    source_counts = df_merged['SOURCE'].value_counts()
    for source, count in source_counts.items():
        print(f"  {source}: {count} rows")

    # Print some statistics
    print("\n" + "-"*80)
    print("COLUMN FILL STATISTICS (Top 20 most populated)")
    print("-"*80)

    fill_stats = []
    for col in df_merged.columns:
        non_null_count = df_merged[col].notna().sum()
        fill_pct = (non_null_count / len(df_merged)) * 100
        fill_stats.append({
            'Column': col,
            'Non-Null Count': non_null_count,
            'Fill %': fill_pct
        })

    fill_df = pd.DataFrame(fill_stats).sort_values('Fill %', ascending=False)
    print(fill_df.head(20).to_string(index=False))

    return df_merged

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == "__main__":
    try:
        # Auto-detect files
        FILES = find_files()

        if FILES is None:
            print("\nERROR: Could not find all required files.")
            print("Please ensure the following files are in the current directory:")
            print("  - Motor KPI*.xlsx")
            print("  - CAM Run Tracker*.xlsx")
            print("  - POG CAM*.xlsx")
            print("  - POG MM*.xlsx")
            print("  - FORMAT GRAL TABLE.xlsx")
            print("  - LISTS_BASIN AND FORM_FAM.xlsx")
            input("\nPress Enter to exit...")
        else:
            # Run the merge
            df_result = merge_all_files(FILES)
            print("\nScript completed successfully!")
            input("\nPress Enter to exit...")
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        input("\nPress Enter to exit...")
